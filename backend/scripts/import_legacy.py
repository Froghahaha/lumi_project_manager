"""
Import projects from the legacy Excel format into the database.

Usage:
  cd backend
  python -m scripts.import_legacy ../docs/项目节点进度表20260508.xlsx [--dry-run]

Excel format (per project, 5 rows covering 5 phases A-R):
  Row 2: 机械设计  Row 3: 生产  Row 4: 调机  Row 5: 验收  Row 6: 尾款
  A: project name (merged)     B: status (merged)     C: equipment (merged)
  D3: salesman   D5: project manager
  F2: start date  F3: contract days  F6: payment progress
  H2: design executor  H3: production executor  H4: tuning executor
  J: phase start  K: planned end  M: actual end
  Q: incidents  R: measures
"""

from __future__ import annotations

import re
import sys
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl.styles.fills import Fill

# Patch broken style parsing in some Excel files
Fill.__init__ = lambda self, *args, **kwargs: None  # type: ignore

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# Add backend to path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db import PRODUCTION_TEMPLATE_ID, engine, init_db  # noqa: E402
from app.models import (  # noqa: E402
    Customer,
    Person,
    PersonRole,
    PhaseIncident,
    PhaseTemplateItem,
    Project,
    ProjectAssignment,
    ProjectEquipment,
    ProjectPhase,
)
from app.utils import hash_password  # noqa: E402
from app.phase_lifecycle import get_init_status  # noqa: E402
from sqlmodel import Session, select  # noqa: E402

# ─── Constants ─────────────────────────────────────────────

PLACEHOLDER_NAMES = {'项目经理', '销售人员', '项目负责人', '/', ''}

PHASE_NAMES = ['机械设计', '生产', '调机', '验收', '尾款']
PHASE_SEQ = {name: i + 1 for i, name in enumerate(PHASE_NAMES)}

# G column -> phase name mapping (column G labels the phase type per row)
G_TO_PHASE: dict[str, str] = {
    '设计': '机械设计',
    '生产': '生产',
    '调机派遣': '调机',
    '验收': '验收',
    '尾款': '尾款',
}

# G column -> executor role
G_TO_ROLE: dict[str, str] = {
    '设计': 'mechanical_designer',
    '生产': 'production_executor',
    '调机派遣': '',  # 此行人不指派 — 验收行的人同时负责调机+验收
    '验收': 'acceptance_executor',
    '尾款': '',
}

# Terminal status when actual_end_date is present (M column has value)
PHASE_COMPLETED_STATUS: dict[int, str] = {
    1: '图纸已下发',   # 机械设计
    2: '生产完成',      # 生产
    3: '安调完成',      # 调机
    4: '已验收',        # 验收
    # 5 尾款 — no status
}

EQUIP_ITEM = re.compile(r'(\d+)\s*台\s*(.+?)(?=\s*\d+\s*台|\s*$|[；;，,\n])')
EQUIP_CLEANUP = re.compile(r'[?？]{2,}.*$')  # garbage like ????????
SKIP_KEYWORDS = {'共', '合计', '总计'}   # summary lines, not real items

DATE_EPOCH = date(1899, 12, 30)

# ─── Helpers ───────────────────────────────────────────────


def parse_date(val: Any) -> date | None:
    """Parse Excel date value."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return DATE_EPOCH + timedelta(days=int(val))
        except (ValueError, OverflowError):
            return None
    if isinstance(val, str) and val.strip():
        val = val.strip()
        if val.startswith('=') or val.startswith('==') or val == '/':
            return None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m-%d', '%m/%d'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def classify_equip(spec: str) -> tuple[str, str]:
    """Classify equipment spec into (category, clean_spec)."""
    spec = spec.strip().rstrip('，。,')
    # Remove garbage suffixes
    spec = EQUIP_CLEANUP.sub('', spec).strip()
    if '关节' in spec:
        return '关节', spec
    if '视觉' in spec:
        return '视觉桁架', spec
    if '桁架' in spec:
        return '桁架', spec
    if '车床' in spec:
        return '联线', spec
    return '其他', spec


def parse_equipment(raw: str | None) -> list[dict]:
    """Parse equipment from legacy Excel column C.

    Examples:
      '4台20KG关节单元' → [{qty:4, cat:'关节', spec:'20KG关节单元'}]
      '1台车床一拖一 1台关节' → [{qty:1, cat:'联线', spec:'车床一拖一'}, {qty:1, cat:'关节', spec:'关节'}]
      '10台视觉\\n21台一拖一' → [{qty:10, cat:'视觉桁架', spec:'视觉'}, {qty:21, cat:'联线', spec:'一拖一'}]
    """
    if not raw:
        return []
    # Normalize: replace newlines with spaces, strip extra spaces
    text = ' '.join(str(raw).replace('\n', ' ').split())
    items: list[dict] = []
    for m in EQUIP_ITEM.finditer(text):
        qty = int(m.group(1))
        spec = m.group(2).strip()
        if not spec:
            continue
        # Skip summary lines like '共53台产品'
        first_word = spec.split()[0] if spec.split() else ''
        if first_word in SKIP_KEYWORDS:
            continue
        cat, clean = classify_equip(spec)
        items.append({'quantity': qty, 'category': cat, 'spec': clean})
    return items


def parse_incidents(raw: str | None, year: int) -> list[dict]:
    """Parse incident text. Multi-incident separated by newlines.
    Incidents start with date prefix like '3-12' or '原因：...' or '现状：...'"""
    if not raw:
        return []
    lines = str(raw).strip().split('\n')
    incidents: list[dict] = []
    current: dict | None = None
    date_prefix = re.compile(r'^(\d{1,2})[-.](\d{1,2})\b')
    tag_prefix = re.compile(r'^(原因|现状|应急|长效)[：:]')

    for line in lines:
        line = line.strip()
        if not line:
            continue
        m_date = date_prefix.match(line)
        m_tag = tag_prefix.match(line)
        if m_date:
            if current:
                incidents.append(current)
            m, d = int(m_date.group(1)), int(m_date.group(2))
            desc = line[m_date.end():].strip()
            current = {
                'occurred_at': date(year, m, d),
                'category': '',
                'description': desc,
            }
        elif m_tag:
            cat = m_tag.group(1)
            desc = line[m_tag.end():].strip()
            cat_map = {'原因': '原因', '现状': '现状', '应急': '应急', '长效': '长效'}
            if current:
                if current['description']:
                    current['description'] += '；' + desc
                else:
                    current['description'] = desc
                current['category'] = cat_map.get(cat, cat)
            else:
                current = {
                    'occurred_at': date(year, 1, 1),
                    'category': cat_map.get(cat, cat),
                    'description': desc,
                }
        else:
            if current:
                if current['description']:
                    current['description'] += '；' + line
                else:
                    current['description'] = line
            else:
                current = {
                    'occurred_at': date(year, 1, 1),
                    'category': '',
                    'description': line,
                }
    if current:
        incidents.append(current)
    return incidents


END_CUSTOMER_RE = re.compile(r'[（(](.+?)[）)]')

ROLE_DEPARTMENT: dict[str, str] = {
    'admin': '管理部',
    'tech_supervisor': '技术部',
    'after_sales_super': '售后部',
    'project_manager': '技术部',
    'salesman': '销售部',
    'sales_assistant': '销售部',
    'mechanical_designer': '技术部',
    'software_designer': '技术部',
    'production_executor': '生产部',
    'tuning_executor': '售后部',
    'acceptance_executor': '售后部',
}


def extract_customer_code(project_name: str) -> str:
    """Extract customer code, stripping terminal customer suffix.
    '宁波国帮-3 （金兔）' → '宁波国帮'"""
    name = str(project_name).strip()
    # Remove terminal customer suffix
    name = END_CUSTOMER_RE.sub('', name).strip()
    if '-' in name:
        return name.rsplit('-', 1)[0]
    return name


def extract_end_customer(project_name: str) -> str | None:
    """Extract terminal customer from parentheses suffix.
    '宁波国帮-3 （金兔）' → '金兔'"""
    m = END_CUSTOMER_RE.search(str(project_name))
    return m.group(1).strip() if m else None

# ─── Main import logic ─────────────────────────────────────


def import_excel(filepath: str, dry_run: bool = False) -> None:
    init_db()

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    max_row = ws.max_row

    # Phase 1: Parse all data
    projects_data: list[dict] = []
    persons_set: dict[str, set[str]] = {}  # name → set of role_codes
    customers_set: set[str] = set()

    # 用合并单元格边界确定每个项目起始行，支持 5/6/7 行不等高项目
    project_bases = sorted(
        mc.min_row for mc in ws.merged_cells.ranges if mc.min_col <= 1 <= mc.max_col
    )

    for base in project_bases:
        name = ws.cell(row=base, column=1).value
        if not name:
            continue
        raw_name = str(name).strip()
        end_customer = extract_end_customer(raw_name)
        # Clean order_no: remove terminal customer suffix
        order_no = END_CUSTOMER_RE.sub('', raw_name).strip()
        customer_code = extract_customer_code(raw_name)
        customers_set.add(customer_code)

        status = ws.cell(row=base, column=2).value
        is_abnormal = str(status).strip() != '正常' if status else False

        equip_raw = ws.cell(row=base, column=3).value
        equipment = parse_equipment(str(equip_raw) if equip_raw else None)

        start_date_raw = ws.cell(row=base, column=6).value
        contract_start = parse_date(start_date_raw)

        days_raw = ws.cell(row=base + 1, column=6).value
        contract_days = None
        if days_raw is not None and not (isinstance(days_raw, str) and days_raw.startswith('=')):
            try:
                contract_days = int(days_raw)
            except (ValueError, TypeError):
                pass

        payment_raw = ws.cell(row=base + 4, column=6).value
        payment = None
        if payment_raw is not None and not (isinstance(payment_raw, str) and payment_raw.startswith('=')):
            try:
                payment = float(payment_raw)
            except (ValueError, TypeError):
                pass

        # F4 = 预计交期, F5 = 实际交期
        expected_delivery = parse_date(ws.cell(row=base + 2, column=6).value)
        actual_delivery_days = None
        ad_raw = ws.cell(row=base + 3, column=6).value
        if ad_raw is not None and not (isinstance(ad_raw, str) and ad_raw.startswith('=')):
            try:
                actual_delivery_days = int(ad_raw)
            except (ValueError, TypeError):
                pass

        # Collect persons from D3(salesman) and D5(PM)
        salesman = ws.cell(row=base + 1, column=4).value
        salesman = None if (not salesman or str(salesman).strip() in PLACEHOLDER_NAMES or str(salesman).startswith('=')) else str(salesman).strip()
        if salesman:
            persons_set.setdefault(salesman, set()).add('salesman')

        pm = ws.cell(row=base + 3, column=4).value
        pm = None if (not pm or str(pm).strip() in PLACEHOLDER_NAMES or str(pm).startswith('=')) else str(pm).strip()
        if pm:
            persons_set.setdefault(pm, set()).add('project_manager')

        # Parse phases from G+H columns (supports multi-phase projects)
        phases: list[dict] = []
        executors: list[dict] = []
        year = contract_start.year if contract_start else 2024
        phase_counters: dict[str, int] = {}

        for i in range(5):
            r = base + i
            g_val = ws.cell(row=r, column=7).value
            if not g_val:
                continue
            g_name = str(g_val).strip()
            phase_name = G_TO_PHASE.get(g_name, g_name)
            if phase_name not in PHASE_SEQ:
                continue

            phase_counters[phase_name] = phase_counters.get(phase_name, 0) + 1
            count = phase_counters[phase_name]
            base_seq = PHASE_SEQ[phase_name]
            seq = base_seq
            sub_name = ''
            if count > 1:
                sub_name = f'{phase_name}{count}'

            # H column: executor
            h_val = ws.cell(row=r, column=8).value
            if h_val:
                h_name = str(h_val).strip()
                if h_name and not h_name.startswith('=') and h_name not in PLACEHOLDER_NAMES:
                    if g_name == '调机派遣':
                        pass  # 调机派遣行的人不指派
                    elif g_name == '验收':
                        # 验收行的人同时负责调机和验收两个工序
                        executors.append({'role_code': 'tuning_executor', 'person_name': h_name, '_row': i, '_match_phase': '调机'})
                        executors.append({'role_code': 'acceptance_executor', 'person_name': h_name, '_row': i})
                        persons_set.setdefault(h_name, set()).update(['tuning_executor', 'acceptance_executor'])
                    elif role_code := G_TO_ROLE.get(g_name, ''):
                        executors.append({'role_code': role_code, 'person_name': h_name, '_row': i})
                        persons_set.setdefault(h_name, set()).add(role_code)
                    else:
                        executors.append({'role_code': '', 'person_name': h_name, '_row': i, 'no_role': True})

            # Dates
            start = parse_date(ws.cell(row=r, column=9).value) or parse_date(ws.cell(row=r, column=10).value)
            planned = parse_date(ws.cell(row=r, column=11).value)
            actual = parse_date(ws.cell(row=r, column=13).value)
            duration_raw = ws.cell(row=r, column=12).value
            planned_duration = None
            if duration_raw is not None and not (isinstance(duration_raw, str) and duration_raw.startswith('=')):
                try:
                    planned_duration = int(duration_raw)
                except (ValueError, TypeError):
                    pass
            incidents_raw = ws.cell(row=r, column=17).value
            incidents = parse_incidents(str(incidents_raw) if incidents_raw else None, year)
            # R column: corrective measures
            measures_raw = ws.cell(row=r, column=18).value
            if measures_raw:
                measures = parse_incidents(str(measures_raw), year)
                for m in measures:
                    m['category'] = '措施'
                incidents.extend(measures)
            # N column: actual_duration
            actual_duration = None
            ad_raw = ws.cell(row=r, column=14).value
            if ad_raw is not None and not (isinstance(ad_raw, str) and ad_raw.startswith('=')):
                try:
                    actual_duration = int(ad_raw)
                except (ValueError, TypeError):
                    pass
            # M column: if actual_end_date present → phase is completed (尾款 always 进行中)
            status = '进行中' if seq == 5 else (PHASE_COMPLETED_STATUS.get(seq, '') if actual else '')
            phases.append({
                '_idx': len(phases),
                'seq': seq,
                'phase_name': phase_name,
                'sub_name': sub_name,
                'status': status,
                'start_date': start,
                'planned_end_date': planned,
                'planned_duration': planned_duration,
                'actual_end_date': actual,
                'actual_duration': actual_duration,
                'incidents': incidents,
            })

        # 生产+安调都有 actual_end_date → 生产应标记为"已发货"
        prod_ph = next((ph for ph in phases if ph['seq'] == 2), None)
        tune_ph = next((ph for ph in phases if ph['seq'] == 3), None)
        if prod_ph and prod_ph['actual_end_date'] and tune_ph and tune_ph['actual_end_date']:
            prod_ph['status'] = '已发货'

        projects_data.append({
            'order_no': order_no,
            'customer_code': customer_code,
            'end_customer': end_customer,
            'is_abnormal': is_abnormal,
            'equipment': equipment,
            'contract_start_date': contract_start,
            'contract_duration_days': contract_days,
            'contract_expected_delivery_date': expected_delivery,
            'contract_actual_delivery_days': actual_delivery_days,
            'contract_payment_progress': payment,
            'phases': phases,
            'assignments': [
                {'role_code': 'salesman', 'person_name': salesman} if salesman else None,
                {'role_code': 'project_manager', 'person_name': pm} if pm else None,
                *executors,
            ],
        })

    # Roles assigned per G column above

    print(f'Parsed {len(projects_data)} projects')
    print(f'Unique customers: {len(customers_set)}')
    print(f'Unique persons: {len(persons_set)}')
    for pname, roles in sorted(persons_set.items()):
        print(f'  {pname}: {roles}')

    if dry_run:
        print('\n[Dry run — no data written]')
        return

    # Phase 2: Write to database
    with Session(engine) as session:
        # Create customers
        customer_map: dict[str, uuid.UUID] = {}
        for code in customers_set:
            existing = session.exec(select(Customer).where(Customer.code == code)).first()
            if existing:
                customer_map[code] = existing.id
                print(f'Customer exists: {code}')
            else:
                c = Customer(code=code, name=code)
                session.add(c)
                session.flush()
                customer_map[code] = c.id
                print(f'Customer created: {code}')

        # Create persons
        person_map: dict[str, uuid.UUID] = {}
        for pname, role_codes in persons_set.items():
            existing = session.exec(select(Person).where(Person.name == pname)).first()
            if existing:
                if not existing.role_code and role_codes:
                    existing.role_code = next(iter(role_codes))
                    session.add(existing)
                # 确保 person_role 同步
                existing_roles = session.exec(
                    select(PersonRole).where(PersonRole.person_id == existing.id)
                ).all()
                if not existing_roles and existing.role_code:
                    session.add(PersonRole(person_id=existing.id, role_code=existing.role_code))
                person_map[pname] = existing.id
                print(f'Person exists: {pname} (role={existing.role_code})')
            else:
                rc = next(iter(role_codes)) if role_codes else ''
                dept = ROLE_DEPARTMENT.get(rc, '')
                p = Person(name=pname, role_code=rc, department=dept,
                           password_hash=hash_password('123456'))
                session.add(p)
                session.flush()
                if rc:
                    session.add(PersonRole(person_id=p.id, role_code=rc))
                person_map[pname] = p.id
                print(f'Person created: {pname} (role={rc})')

        session.commit()

        # Create projects
        template_items = list(session.exec(
            select(PhaseTemplateItem).where(PhaseTemplateItem.template_id == PRODUCTION_TEMPLATE_ID)
        ))

        created = 0
        skipped = 0
        for pd in projects_data:
            existing_proj = session.exec(
                select(Project).where(Project.order_no == pd['order_no'])
            ).first()
            if existing_proj:
                skipped += 1
                continue

            cust_id = customer_map.get(pd['customer_code'])
            if not cust_id:
                print(f'  WARNING: customer not found for {pd["order_no"]}')
                continue

            proj = Project(
                order_no=pd['order_no'],
                customer_id=cust_id,
                end_customer=pd.get('end_customer'),
                template_id=PRODUCTION_TEMPLATE_ID,
                contract_start_date=pd['contract_start_date'],
                contract_duration_days=pd['contract_duration_days'],
                contract_expected_delivery_date=pd.get('contract_expected_delivery_date'),
                contract_actual_delivery_days=pd.get('contract_actual_delivery_days'),
                contract_payment_progress=pd['contract_payment_progress'],
                is_abnormal=pd['is_abnormal'],
            )
            session.add(proj)
            session.flush()

            # Equipment
            for eq in pd['equipment']:
                session.add(ProjectEquipment(
                    project_id=proj.id,
                    category=eq['category'],
                    spec=eq['spec'],
                    quantity=eq['quantity'],
                ))

            # Phases + Incidents
            phases_by_row: dict[int, uuid.UUID] = {}  # _row -> phase_id
            for ph_data in pd['phases']:
                tmpl_item = next(
                    (ti for ti in template_items if ti.seq == ph_data['seq']), None
                )
                ph = ProjectPhase(
                    project_id=proj.id,
                    seq=ph_data['seq'],
                    phase_name=ph_data['phase_name'],
                    sub_name=ph_data.get('sub_name', ''),
                    status=ph_data.get('status', ''),
                    start_date=ph_data['start_date'],
                    planned_end_date=ph_data['planned_end_date'],
                    planned_duration=ph_data.get('planned_duration'),
                    actual_end_date=ph_data['actual_end_date'],
                    actual_duration=ph_data.get('actual_duration'),
                )
                session.add(ph)
                session.flush()
                phases_by_row[ph_data['_idx']] = ph.id

                for inc in ph_data['incidents']:
                    session.add(PhaseIncident(
                        phase_id=ph.id,
                        occurred_at=inc['occurred_at'],
                        category=inc['category'],
                        description=inc['description'],
                    ))

            # Assignments: phase executors bind to their specific phase;
            # global roles (salesman, project_manager) have phase_id=None
            for a in pd['assignments']:
                if not a:
                    continue
                pname = a['person_name']
                role_code = a.get('role_code', '')
                if not pname or pname not in person_map:
                    continue
                if a.get('no_role'):
                    continue
                # Match by _row; for _match_phase, find the row of that phase type
                row = a.get('_row')
                if a.get('_match_phase'):
                    # Find the row index where this phase type was parsed
                    target = a['_match_phase']
                    for ph_data in pd['phases']:
                        if ph_data['phase_name'] == target:
                            row = ph_data['_idx']
                            break
                ph_id = phases_by_row.get(row) if row is not None else None
                if ph_id or row is None:
                    session.add(ProjectAssignment(
                        project_id=proj.id,
                        person_name=pname,
                        role_code=role_code,
                        phase_id=ph_id,
                    ))

            created += 1

        session.commit()
        print(f'\nProjects: {created} created, {skipped} skipped (already exist)')


if __name__ == '__main__':
    dry = '--dry-run' in sys.argv
    path = sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith('--') else None
    if not path:
        print('Usage: python -m scripts.import_legacy <excel_file> [--dry-run]')
        sys.exit(1)
    import_excel(path, dry_run=dry)
